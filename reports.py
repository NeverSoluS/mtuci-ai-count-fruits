"""
Генерация отчётов в PDF и Excel форматах.
- PDF: сводная статистика + детальная история
- Excel: два листа (История + Сводка) с красивым форматированием
"""

import os
import json
from datetime import datetime
from collections import defaultdict

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config


 
# Шрифт для кириллицы (поиск в стандартных местах)
FONT_NAME = 'Helvetica'
FONT_PATHS = [
    'C:/Windows/Fonts/arial.ttf',
    'C:/Windows/Fonts/Arial.ttf',
    'arial.ttf',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    '/System/Library/Fonts/Helvetica.ttc',
]

for fp in FONT_PATHS:
    if os.path.exists(fp):
        try:
            pdfmetrics.registerFont(TTFont('CyrillicFont', fp))
            FONT_NAME = 'CyrillicFont'
            print(f"[REPORTS] Найден шрифт с кириллицей: {fp}")
            break
        except Exception as e:
            print(f"[REPORTS] Не удалось загрузить {fp}: {e}")


# Словари для отображения режимов и движков в человекочитаемом виде
MODE_NAMES = {
    'detection': 'Детекция',
    'segmentation': 'Сегментация',
    'classification': 'Классификация',
}

ENGINE_NAMES = {
    'yolo': 'YOLOv8m',
    'owlvit': 'OWLv2',
    'owl-vit': 'OWL-ViT',
}


 
# Вспомогательные функции для обработки истории запросов
def _safe_by_class(record):
    """Безопасно извлечь by_class из записи."""
    bc = record.get('by_class', {})
    if isinstance(bc, str):
        try:
            bc = json.loads(bc)
        except Exception:
            bc = {}
    return bc if isinstance(bc, dict) else {}


def _compute_summary(history):
    """Вычислить сводную статистику по истории."""
    total_fruits = 0
    by_class_total = defaultdict(int)
    by_mode = defaultdict(int)
    by_engine = defaultdict(int)
    
    for record in history:
        total = record.get('total_fruits', 0) or 0
        total_fruits += total
        
        bc = _safe_by_class(record)
        for cls, count in bc.items():
            try:
                by_class_total[cls] += int(count)
            except (ValueError, TypeError):
                pass
        
        mode = record.get('mode', 'detection')
        by_mode[mode] += 1
        
        engine = record.get('engine', 'yolo')
        by_engine[engine] += 1
    
    return {
        'total_requests': len(history),
        'total_fruits': total_fruits,
        'by_class': dict(sorted(by_class_total.items(), key=lambda x: -x[1])),
        'by_mode': dict(by_mode),
        'by_engine': dict(by_engine),
    }

# PDF генерация
def generate_pdf(history, output_path=None):
    """
    Генерация PDF-отчёта с датой в имени файла.
    
    Содержит:
    - Первую страницу со сводной статистикой
    - Детальную историю запросов
    """
    # Имя файла с датой
    if output_path is None:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        output_path = os.path.join(config.REPORTS_DIR, f'report_{timestamp}.pdf')
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    c = canvas.Canvas(output_path, pagesize=A4)
    w, h = A4
    
    # Первый лист: сводная статистика 
    c.setFont(FONT_NAME, 18)
    c.drawString(20*mm, h - 20*mm, "Отчёт: Подсчёт фруктов")

    c.setFont(FONT_NAME, 11)
    c.drawString(20*mm, h - 30*mm,
                 f"Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    summary = _compute_summary(history)
    
    # Общая статистика
    c.setFont(FONT_NAME, 14)
    c.drawString(20*mm, h - 45*mm, "Общая статистика:")
    
    c.setFont(FONT_NAME, 11)
    y = h - 55*mm
    c.drawString(25*mm, y, f"• Всего запросов: {summary['total_requests']}")
    y -= 7*mm
    c.drawString(25*mm, y, f"• Всего фруктов: {summary['total_fruits']}")
    y -= 7*mm
    
    # По режимам
    mode_str = ', '.join(
        f"{MODE_NAMES.get(m, m)}: {cnt}" 
        for m, cnt in summary['by_mode'].items()
    )
    c.drawString(25*mm, y, f"• По режимам: {mode_str or '-'}")
    y -= 7*mm
    
    # По движкам
    engine_str = ', '.join(
        f"{ENGINE_NAMES.get(e, e)}: {cnt}" 
        for e, cnt in summary['by_engine'].items()
    )
    c.drawString(25*mm, y, f"• По движкам: {engine_str or '-'}")
    y -= 14*mm
    
    # Топ-10 фруктов
    c.setFont(FONT_NAME, 14)
    c.drawString(20*mm, y, "Топ-10 фруктов:")
    y -= 9*mm
    c.setFont(FONT_NAME, 11)
    
    top_classes = list(summary['by_class'].items())[:10]
    for cls, count in top_classes:
        c.drawString(25*mm, y, f"• {cls}: {count}")
        y -= 6*mm
    
    if not top_classes:
        c.drawString(25*mm, y, "• (нет данных)")
    
    # Новая страница для детальной истории 
    c.showPage()
    
    c.setFont(FONT_NAME, 16)
    c.drawString(20*mm, h - 20*mm, "Детальная история запросов")
    
    c.setFont(FONT_NAME, 11)
    c.drawString(20*mm, h - 28*mm, f"Всего записей: {len(history)}")

    # Заголовок таблицы
    y = h - 40*mm
    c.setFont(FONT_NAME, 10)
    c.setFillColorRGB(0.2, 0.2, 0.2)
    c.drawString(15*mm, y, "Время")
    c.drawString(55*mm, y, "Файл")
    c.drawString(95*mm, y, "Режим")
    c.drawString(120*mm, y, "Движок")
    c.drawString(148*mm, y, "Всего")
    
    # Линия-разделитель
    c.setStrokeColorRGB(0.5, 0.5, 0.5)
    c.line(15*mm, y - 1*mm, w - 15*mm, y - 1*mm)
    c.setStrokeColorRGB(0, 0, 0)
    c.setFillColorRGB(0, 0, 0)
    
    y -= 8*mm
    
    # Данные
    for record in history[:100]:  # максимум 100 записей
        if y < 30*mm:
            c.showPage()
            y = h - 20*mm
        
        c.setFont(FONT_NAME, 9)
        
        timestamp = record.get('timestamp', '')[:19]
        filename = (record.get('filename', '') or '')[:22]
        mode = MODE_NAMES.get(record.get('mode', 'detection'), '?')
        engine = ENGINE_NAMES.get(record.get('engine', 'yolo'), 
                                   record.get('engine', '?'))
        total = record.get('total_fruits', 0)
        
        c.drawString(15*mm, y, timestamp)
        c.drawString(55*mm, y, filename)
        c.drawString(95*mm, y, mode)
        c.drawString(120*mm, y, engine)
        c.drawString(148*mm, y, str(total))
        
        # Детали (серым цветом)
        bc = _safe_by_class(record)
        details = ', '.join(f"{k}:{v}" for k, v in bc.items()) if bc else '-'
        if len(details) > 80:
            details = details[:77] + '...'
        
        c.setFont(FONT_NAME, 7)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(17*mm, y - 3.5*mm, details)
        c.setFillColorRGB(0, 0, 0)
        
        y -= 10*mm
    
    c.save()
    print(f"[REPORTS] ✓ PDF сохранён: {output_path}")
    return output_path

# Генерация Excel отчёта
def generate_excel(history, output_path=None):
    """
    Генерация Excel-отчёта с датой в имени файла.
    
    Содержит два листа:
    - "История" - детальная таблица
    - "Сводка" - агрегированная статистика
    """
    # Имя файла с датой
    if output_path is None:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        output_path = os.path.join(config.REPORTS_DIR, f'report_{timestamp}.xlsx')
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    
    # Стили  
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", 
                               fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    
   
    # Лист 1: История запросов
    ws = wb.active
    ws.title = "История"
    
    headers = ['ID', 'Время', 'Файл', 'Режим', 'Движок', 'Всего', 'Детали']
    ws.append(headers)
    
    # Стилизация заголовков
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Заполнение данных
    for record in history:
        bc = _safe_by_class(record)
        details = ', '.join(f"{k}: {v}" for k, v in bc.items()) if bc else '-'
        mode = MODE_NAMES.get(record.get('mode', 'detection'), 
                               record.get('mode', ''))
        engine = ENGINE_NAMES.get(record.get('engine', 'yolo'), 
                                   record.get('engine', ''))
        
        ws.append([
            record.get('id', ''),
            record.get('timestamp', ''),
            record.get('filename', ''),
            mode,
            engine,
            record.get('total_fruits', 0),
            details,
        ])
    
    # Рамки для всех ячеек
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Ширина колонок
    column_widths = {
        'A': 6,   # ID
        'B': 22,  # Время
        'C': 35,  # Файл
        'D': 14,  # Режим
        'E': 12,  # Движок
        'F': 8,   # Всего
        'G': 60,  # Детали
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Закрепить шапку
    ws.freeze_panes = 'A2'
    
   
    # Лист 2: Сводная статистика
    ws2 = wb.create_sheet("Сводка")
    summary = _compute_summary(history)
    
    # Общая статистика
    ws2.append(['Общая статистика'])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14, color="0071E3")
    
    ws2.append(['Показатель', 'Значение'])
    for cell in ws2[2]:
        cell.font = header_font
        cell.fill = header_fill
    
    ws2.append(['Всего запросов', summary['total_requests']])
    ws2.append(['Всего фруктов', summary['total_fruits']])
    ws2.append([])
    
    # По режимам
    ws2.append(['Статистика по режимам'])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=12)
    ws2.append(['Режим', 'Количество запросов'])
    for cell in ws2[ws2.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for mode, cnt in summary['by_mode'].items():
        ws2.append([MODE_NAMES.get(mode, mode), cnt])
    ws2.append([])
    
    # По движкам
    ws2.append(['Статистика по движкам'])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=12)
    ws2.append(['Движок', 'Количество запросов'])
    for cell in ws2[ws2.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for engine, cnt in summary['by_engine'].items():
        ws2.append([ENGINE_NAMES.get(engine, engine), cnt])
    ws2.append([])
    
    # Топ фруктов
    ws2.append(['Статистика по фруктам (все запросы)'])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=12)
    ws2.append(['Фрукт', 'Всего обнаружено'])
    for cell in ws2[ws2.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for cls, cnt in summary['by_class'].items():
        ws2.append([cls, cnt])
    
    # Ширина колонок
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 25
    
    # Сохранение
    wb.save(output_path)
    print(f"[REPORTS] ✓ Excel сохранён: {output_path}")
    return output_path